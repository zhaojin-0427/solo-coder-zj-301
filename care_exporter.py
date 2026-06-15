import pandas as pd
import numpy as np
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def export_care_report_to_excel(
    care_df, consistency, routine_result, nw_diff_result,
    deviation_result, patterns, care_summary, filters_info=None
):
    output = BytesIO()
    wb = openpyxl.Workbook()

    header_font = Font(name='Arial Unicode MS', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    section_font = Font(name='Arial Unicode MS', bold=True, size=11, color='6366F1')
    sub_section_font = Font(name='Arial Unicode MS', bold=True, size=10, color='374151')
    normal_font = Font(name='Arial Unicode MS', size=10)
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    warning_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    info_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    success_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')

    def style_header_row(ws, row, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

    def style_data_cell(ws, row, col, value, fill=None):
        if value is None:
            value = ''
        elif isinstance(value, (float, np.floating)):
            if pd.isna(value) or np.isnan(value) or np.isinf(value):
                value = ''
        elif pd.isna(value):
            value = ''
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = normal_font
        cell.alignment = wrap_alignment
        cell.border = thin_border
        if fill:
            cell.fill = fill
        return cell

    ws1 = wb.active
    ws1.title = '照护记录明细'
    ws1.cell(row=1, column=1, value='照护协同交接报告 - 照护记录明细').font = Font(
        name='Arial Unicode MS', bold=True, size=16, color='6366F1'
    )
    ws1.merge_cells('A1:I1')
    ws1.cell(row=2, column=1, value=f'生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}').font = normal_font
    r = 4
    detail_headers = ['日期', '照护人', '睡前流程执行', '完成率(%)', '安抚方式',
                       '夜醒响应方式', '环境变动', '临时事件', '交接备注']
    for j, h in enumerate(detail_headers):
        style_data_cell(ws1, r, j + 1, h)
    style_header_row(ws1, r, len(detail_headers))
    r += 1
    if care_df is not None and len(care_df) > 0:
        df_sorted = care_df.sort_values('date').reset_index(drop=True)
        for _, row_data in df_sorted.iterrows():
            date_val = row_data['date'].strftime('%Y-%m-%d') if pd.notna(row_data['date']) else ''
            style_data_cell(ws1, r, 1, date_val)
            style_data_cell(ws1, r, 2, row_data.get('caregiver', ''))
            routine_str = ','.join(row_data.get('routine_items', [])) if isinstance(row_data.get('routine_items'), list) else str(row_data.get('bedtime_routine', ''))
            style_data_cell(ws1, r, 3, routine_str)
            rate = row_data.get('routine_completion_rate', 0)
            rate_fill = None
            if isinstance(rate, (int, float)):
                rate_fill = success_fill if rate >= 60 else (warning_fill if rate < 40 else None)
            style_data_cell(ws1, r, 4, rate, fill=rate_fill)
            style_data_cell(ws1, r, 5, row_data.get('soothing_method', ''))
            style_data_cell(ws1, r, 6, row_data.get('nw_response', ''))
            style_data_cell(ws1, r, 7, row_data.get('env_change', ''))
            style_data_cell(ws1, r, 8, row_data.get('temp_event', ''))
            style_data_cell(ws1, r, 9, row_data.get('handover_note', ''))
            r += 1
    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 10
    ws1.column_dimensions['C'].width = 35
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 14
    ws1.column_dimensions['F'].width = 18
    ws1.column_dimensions['G'].width = 14
    ws1.column_dimensions['H'].width = 20
    ws1.column_dimensions['I'].width = 25

    ws2 = wb.create_sheet('交接一致性评分')
    r2 = 1
    ws2.cell(row=r2, column=1, value='照护交接一致性评分').font = Font(
        name='Arial Unicode MS', bold=True, size=14, color='6366F1'
    )
    ws2.merge_cells('A1:D1')
    r2 += 2
    score_items = [
        ('综合评分', f"{consistency.get('overall', 0)} 分", consistency.get('level', '')),
        ('睡前流程一致性', f"{consistency.get('routine_score', 0)} 分", ''),
        ('夜醒响应一致性', f"{consistency.get('response_score', 0)} 分", ''),
        ('交接充分性', f"{consistency.get('handover_score', 0)} 分", ''),
        ('评语', consistency.get('details', ''), ''),
    ]
    headers2 = ['指标', '得分/内容', '说明']
    for j, h in enumerate(headers2):
        style_data_cell(ws2, r2, j + 1, h)
    style_header_row(ws2, r2, len(headers2))
    r2 += 1
    for label, value, note in score_items:
        style_data_cell(ws2, r2, 1, label)
        style_data_cell(ws2, r2, 2, value)
        style_data_cell(ws2, r2, 3, note)
        r2 += 1
    if filters_info:
        r2 += 2
        ws2.cell(row=r2, column=1, value='筛选条件').font = section_font
        r2 += 1
        for k, v in filters_info.items():
            style_data_cell(ws2, r2, 1, k)
            style_data_cell(ws2, r2, 2, str(v))
            ws2.merge_cells(start_row=r2, start_column=2, end_row=r2, end_column=4)
            r2 += 1
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 30
    ws2.column_dimensions['D'].width = 20

    ws3 = wb.create_sheet('睡前流程执行情况')
    r3 = 1
    ws3.cell(row=r3, column=1, value='睡前流程执行情况').font = Font(
        name='Arial Unicode MS', bold=True, size=14, color='6366F1'
    )
    ws3.merge_cells('A1:D1')
    r3 += 2
    ws3.cell(row=r3, column=1, value='一、总体完成率').font = section_font
    r3 += 1
    style_data_cell(ws3, r3, 1, '平均完成率')
    style_data_cell(ws3, r3, 2, f"{routine_result.get('overall_rate', 0)}%")
    r3 += 2
    ws3.cell(row=r3, column=1, value='二、各照护人完成率').font = section_font
    r3 += 1
    cg_headers = ['照护人', '平均完成率(%)', '最低(%)', '最高(%)', '记录数']
    for j, h in enumerate(cg_headers):
        style_data_cell(ws3, r3, j + 1, h)
    style_header_row(ws3, r3, len(cg_headers))
    r3 += 1
    for cg, info in routine_result.get('by_caregiver', {}).items():
        style_data_cell(ws3, r3, 1, cg)
        style_data_cell(ws3, r3, 2, info.get('avg_rate', 0))
        style_data_cell(ws3, r3, 3, info.get('min_rate', 0))
        style_data_cell(ws3, r3, 4, info.get('max_rate', 0))
        style_data_cell(ws3, r3, 5, info.get('count', 0))
        r3 += 1
    r3 += 1
    ws3.cell(row=r3, column=1, value='三、各项目执行频率').font = section_font
    r3 += 1
    item_headers = ['流程项目', '执行频率(%)']
    for j, h in enumerate(item_headers):
        style_data_cell(ws3, r3, j + 1, h)
    style_header_row(ws3, r3, len(item_headers))
    r3 += 1
    item_freq = routine_result.get('item_frequency', {})
    for item, freq in sorted(item_freq.items(), key=lambda x: x[1], reverse=True):
        fill = success_fill if freq >= 60 else (warning_fill if freq < 40 else None)
        style_data_cell(ws3, r3, 1, item)
        style_data_cell(ws3, r3, 2, freq, fill=fill)
        r3 += 1
    ws3.column_dimensions['A'].width = 22
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 14
    ws3.column_dimensions['D'].width = 14
    ws3.column_dimensions['E'].width = 12

    ws4 = wb.create_sheet('照护人差异分析')
    r4 = 1
    ws4.cell(row=r4, column=1, value='照护人差异分析').font = Font(
        name='Arial Unicode MS', bold=True, size=14, color='6366F1'
    )
    ws4.merge_cells('A1:E1')
    r4 += 2
    ws4.cell(row=r4, column=1, value='一、各照护人核心指标').font = section_font
    r4 += 1
    cg_detail_headers = ['照护人', '记录数', '平均夜醒(次)', '主要安抚方式', '主要夜醒响应']
    for j, h in enumerate(cg_detail_headers):
        style_data_cell(ws4, r4, j + 1, h)
    style_header_row(ws4, r4, len(cg_detail_headers))
    r4 += 1
    for cg, info in nw_diff_result.get('by_caregiver', {}).items():
        style_data_cell(ws4, r4, 1, cg)
        style_data_cell(ws4, r4, 2, info.get('count', 0))
        avg_nw = info.get('avg_nightwakings')
        style_data_cell(ws4, r4, 3, avg_nw if avg_nw is not None else '无数据')
        style_data_cell(ws4, r4, 4, info.get('primary_soothing', ''))
        style_data_cell(ws4, r4, 5, info.get('primary_response', ''))
        r4 += 1
    r4 += 1
    ws4.cell(row=r4, column=1, value='二、夜醒响应方式分布').font = section_font
    r4 += 1
    for cg, info in nw_diff_result.get('by_caregiver', {}).items():
        ws4.cell(row=r4, column=1, value=cg).font = sub_section_font
        r4 += 1
        resp_dist = info.get('nw_response_distribution', {})
        for resp, count in resp_dist.items():
            style_data_cell(ws4, r4, 1, resp)
            style_data_cell(ws4, r4, 2, count)
            r4 += 1
        r4 += 1
    ws4.column_dimensions['A'].width = 20
    ws4.column_dimensions['B'].width = 14
    ws4.column_dimensions['C'].width = 18
    ws4.column_dimensions['D'].width = 20
    ws4.column_dimensions['E'].width = 22

    ws5 = wb.create_sheet('干预偏差对比')
    r5 = 1
    ws5.cell(row=r5, column=1, value='干预计划执行偏差对比').font = Font(
        name='Arial Unicode MS', bold=True, size=14, color='6366F1'
    )
    ws5.merge_cells('A1:E1')
    r5 += 2
    style_data_cell(ws5, r5, 1, '执行一致性评分')
    dev_score = deviation_result.get('deviation_score', 0)
    dev_fill = success_fill if dev_score >= 70 else (warning_fill if dev_score < 50 else None)
    style_data_cell(ws5, r5, 2, f"{dev_score} 分", fill=dev_fill)
    r5 += 2
    ws5.cell(row=r5, column=1, value='执行检查明细').font = section_font
    r5 += 1
    dev_headers = ['日期', '照护人', '入睡窗口', '流程完成', '推荐响应']
    for j, h in enumerate(dev_headers):
        style_data_cell(ws5, r5, j + 1, h)
    style_header_row(ws5, r5, len(dev_headers))
    r5 += 1
    for d in deviation_result.get('deviation_details', [])[:100]:
        date_str = d['date'].strftime('%Y-%m-%d') if hasattr(d['date'], 'strftime') else str(d['date'])
        style_data_cell(ws5, r5, 1, date_str)
        style_data_cell(ws5, r5, 2, d.get('caregiver', ''))
        checks = d.get('checks', {})
        style_data_cell(ws5, r5, 3, '✓' if checks.get('bedtime_in_window', True) else '✗',
                        fill=success_fill if checks.get('bedtime_in_window', True) else warning_fill)
        style_data_cell(ws5, r5, 4, '✓' if checks.get('routine_completed', True) else '✗',
                        fill=success_fill if checks.get('routine_completed', True) else warning_fill)
        style_data_cell(ws5, r5, 5, '✓' if checks.get('recommended_response', True) else '✗',
                        fill=success_fill if checks.get('recommended_response', True) else warning_fill)
        r5 += 1
    ws5.column_dimensions['A'].width = 14
    ws5.column_dimensions['B'].width = 12
    ws5.column_dimensions['C'].width = 14
    ws5.column_dimensions['D'].width = 14
    ws5.column_dimensions['E'].width = 14

    ws6 = wb.create_sheet('风险提醒与建议')
    r6 = 1
    ws6.cell(row=r6, column=1, value='风险提醒与下一步协同建议').font = Font(
        name='Arial Unicode MS', bold=True, size=14, color='6366F1'
    )
    ws6.merge_cells('A1:C1')
    r6 += 2
    ws6.cell(row=r6, column=1, value='一、自动识别模式').font = section_font
    r6 += 1
    pattern_fills = {
        'warning': warning_fill,
        'info': info_fill,
        'success': success_fill,
    }
    for p in patterns:
        p_fill = pattern_fills.get(p.get('type', ''), None)
        p_type_label = {'warning': '⚠️ 风险', 'info': 'ℹ️ 提示', 'success': '✅ 良好'}.get(p.get('type', ''), '')
        style_data_cell(ws6, r6, 1, f"{p_type_label}：{p.get('title', '')}", fill=p_fill)
        ws6.merge_cells(start_row=r6, start_column=1, end_row=r6, end_column=3)
        r6 += 1
        style_data_cell(ws6, r6, 1, p.get('detail', ''))
        ws6.merge_cells(start_row=r6, start_column=1, end_row=r6, end_column=3)
        r6 += 1
    r6 += 1
    ws6.cell(row=r6, column=1, value='二、照护人差异提醒').font = section_font
    r6 += 1
    for alert in nw_diff_result.get('difference_alerts', []):
        alert_fill = warning_fill if alert.get('type') == 'warning' else info_fill
        style_data_cell(ws6, r6, 1, alert.get('title', ''), fill=alert_fill)
        ws6.merge_cells(start_row=r6, start_column=1, end_row=r6, end_column=3)
        r6 += 1
        style_data_cell(ws6, r6, 1, alert.get('detail', ''))
        ws6.merge_cells(start_row=r6, start_column=1, end_row=r6, end_column=3)
        r6 += 1
    r6 += 1
    ws6.cell(row=r6, column=1, value='三、干预偏差提醒').font = section_font
    r6 += 1
    for alert in deviation_result.get('alerts', []):
        alert_fill = warning_fill if alert.get('type') == 'warning' else info_fill
        style_data_cell(ws6, r6, 1, alert.get('title', ''), fill=alert_fill)
        ws6.merge_cells(start_row=r6, start_column=1, end_row=r6, end_column=3)
        r6 += 1
        style_data_cell(ws6, r6, 1, alert.get('detail', ''))
        ws6.merge_cells(start_row=r6, start_column=1, end_row=r6, end_column=3)
        r6 += 1
    r6 += 1
    ws6.cell(row=r6, column=1, value='四、下一步协同建议').font = section_font
    r6 += 1
    suggestions = _generate_suggestions(consistency, routine_result, nw_diff_result, deviation_result, patterns)
    for i, sug in enumerate(suggestions, 1):
        style_data_cell(ws6, r6, 1, f'{i}. {sug}')
        ws6.merge_cells(start_row=r6, start_column=1, end_row=r6, end_column=3)
        r6 += 1
    ws6.column_dimensions['A'].width = 50
    ws6.column_dimensions['B'].width = 30
    ws6.column_dimensions['C'].width = 30

    wb.save(output)
    output.seek(0)
    return output


def _generate_suggestions(consistency, routine_result, nw_diff_result, deviation_result, patterns):
    suggestions = []
    if consistency.get('routine_score', 0) < 60:
        suggestions.append('统一各照护人的睡前流程执行标准，确保完成率≥60%')
    if consistency.get('response_score', 0) < 60:
        suggestions.append('协商统一的夜醒响应策略，避免不同照护人方式差异过大')
    if consistency.get('handover_score', 0) < 50:
        suggestions.append('加强照护人交接记录，每日填写交接备注，确保信息传递完整')
    if routine_result.get('overall_rate', 0) < 50:
        suggestions.append('睡前流程完成率偏低，建议精简流程至核心5项，确保可执行性')
    by_cg = nw_diff_result.get('by_caregiver', {})
    if len(by_cg) >= 2:
        nw_vals = {cg: info.get('avg_nightwakings') for cg, info in by_cg.items() if info.get('avg_nightwakings') is not None}
        if len(nw_vals) >= 2:
            max_cg = max(nw_vals, key=nw_vals.get)
            min_cg = min(nw_vals, key=nw_vals.get)
            if nw_vals[max_cg] - nw_vals[min_cg] >= 1.0:
                suggestions.append(f'重点培训{max_cg}的安抚技巧，参考{min_cg}的有效做法')
    if deviation_result.get('deviation_score', 0) < 60:
        suggestions.append('干预计划执行偏差较大，建议每日对照检查清单，确保照护人知悉执行要求')
    for p in patterns:
        if p.get('type') == 'warning' and '睡前流程缺项' in p.get('title', ''):
            suggestions.append('睡前流程缺项直接关联夜醒增加，即使时间紧张也应保证核心3项执行')
    if not suggestions:
        suggestions.append('当前照护协作良好，继续保持现有策略并定期沟通')
        suggestions.append('建议每周回顾照护交接记录，确保信息一致性')
    return suggestions
