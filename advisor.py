import pandas as pd
import numpy as np
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from data_processor import minutes_to_time_display
from analyzer import (
    compute_basic_stats, compute_stability_score, detect_patterns,
    compare_to_norms, age_based_sleep_norms
)


def generate_sleep_advice(df, patterns, stability, stats):
    advice = {
        'bedtime_window': {},
        'nap_adjustment': [],
        'general_tips': [],
        'priority_actions': []
    }
    
    if len(df) < 3:
        advice['general_tips'].append('数据记录不足3天，建议持续记录至少7天以获得更准确的节律建议。')
        return advice
    
    recent = df.tail(min(14, len(df)))
    
    bt_mean = recent['bedtime_minutes'].mean()
    bt_std = recent['bedtime_minutes'].std() if len(recent) > 1 else 30
    
    if pd.notna(bt_mean):
        window_start = minutes_to_time_display(bt_mean - 30)
        window_end = minutes_to_time_display(bt_mean + 15)
        ideal_start = minutes_to_time_display(-150)
        ideal_end = minutes_to_time_display(-90)
        
        if bt_std > 60:
            consistency = f"⚠️ 当前入睡波动较大(±{bt_std:.0f}分钟)，建议将波动控制在±30分钟内"
        elif bt_std > 30:
            consistency = f"📊 入睡波动一般(±{bt_std:.0f}分钟)，建议进一步固定时间"
        else:
            consistency = f"✅ 入睡时间稳定(±{bt_std:.0f}分钟)，继续保持！"
        
        advice['bedtime_window'] = {
            'current': {
                'avg': minutes_to_time_display(bt_mean),
                'window': f"{window_start} ~ {window_end}",
                'consistency': consistency
            },
            'recommended': {
                'window': f"{ideal_start} ~ {ideal_end}",
                'reason': '多数儿科研究建议婴幼儿在21:30-22:30间入睡，此时段皮质醇下降、褪黑素上升'
            }
        }
    
    age_group = df['age_group'].mode().iloc[0] if df['age_group'].notna().any() else '未知'
    norms = age_based_sleep_norms()
    
    if age_group in norms:
        n = norms[age_group]
        current_naps = recent['naps_count'].mean()
        current_nap_minutes = recent['total_nap_minutes'].mean()
        
        advice['nap_adjustment'].append(
            f"📊 本月龄({age_group})推荐小睡 {n['naps'][0]}-{n['naps'][1]} 次，总时长约3-5小时"
        )
        
        if current_naps > n['naps'][1] + 0.5:
            advice['nap_adjustment'].append(
                f"🔽 当前平均{current_naps:.1f}次小睡，偏多。建议减少一次小睡，适当延长清醒间隔"
            )
            advice['priority_actions'].append({
                'level': 'high',
                'action': '减少白天小睡次数',
                'detail': f'从{current_naps:.1f}次逐步减到{n["naps"][1]}次，避免碎片化睡眠'
            })
        elif current_naps < n['naps'][0] - 0.5:
            advice['nap_adjustment'].append(
                f"🔼 当前平均{current_naps:.1f}次小睡，偏少。建议增加一次小憩，避免过度疲劳"
            )
            advice['priority_actions'].append({
                'level': 'high',
                'action': '增加白天小睡',
                'detail': '过度疲劳反而会导致夜醒增加，注意观察揉眼、打哈欠等困倦信号'
            })
        
        last_nap_valid = recent.dropna(subset=['last_nap_minutes'])
        if len(last_nap_valid) >= 3:
            avg_last_nap = last_nap_valid['last_nap_minutes'].mean()
            if avg_last_nap >= 15 * 60:
                advice['nap_adjustment'].append(
                    f"⚠️ 最后一觉平均{minutes_to_time_display(avg_last_nap)}结束，偏晚。建议15:00前结束白天最后一觉"
                )
                advice['priority_actions'].append({
                    'level': 'medium',
                    'action': '提前最后一觉结束时间',
                    'detail': '将最后一觉限制在15:00前结束，给睡前足够的清醒时间积累睡眠压力'
                })
        
        if current_nap_minutes / 60 > 5.5:
            advice['nap_adjustment'].append(
                f"⚠️ 白天总小睡{current_nap_minutes/60:.1f}小时过长，可能影响夜间连续睡眠"
            )
        elif current_nap_minutes / 60 < 2 and current_naps >= 1:
            advice['nap_adjustment'].append(
                f"🔼 白天总小睡{current_nap_minutes/60:.1f}小时偏短，尝试延长每次小睡30分钟"
            )
    
    nw_avg = recent['nightwakings'].mean()
    if nw_avg > 2:
        advice['priority_actions'].append({
            'level': 'high',
            'action': '排查夜醒原因并建立安抚程序',
            'detail': f'近期平均夜醒{nw_avg:.1f}次，建议先建立固定的睡前程序（洗澡-抚触-喂奶-故事），再逐步拉长安抚间隔'
        })
    
    norm_comparison = compare_to_norms(
        age_group,
        stats.get('avg_total_sleep_hours'),
        stats.get('avg_night_sleep_hours'),
        stats.get('avg_naps_count')
    )
    for label, status, detail in norm_comparison:
        if status == '偏低':
            advice['general_tips'].append(f"⚠️ {detail}")
        elif status == '偏高':
            advice['general_tips'].append(f"💡 {detail}，注意观察是否影响其他作息")
        else:
            advice['general_tips'].append(f"✅ {detail}")
    
    milk_valid = recent['milk_amount_ml'].dropna()
    if len(milk_valid) >= 3:
        avg_milk = milk_valid.mean()
        if avg_milk < 600:
            advice['priority_actions'].append({
                'level': 'medium',
                'action': '增加白天奶量摄入',
                'detail': f'近期日均{avg_milk:.0f}ml偏低，可能导致夜醒求奶。建议白天增加1-2次喂奶量'
            })
    
    if stability.get('overall', 50) < 60:
        advice['priority_actions'].append({
            'level': 'high',
            'action': '建立固定的作息时间表',
            'detail': f'当前稳定度{stability["overall"]:.0f}分偏低。固定起床、喂奶、小睡、入睡时间，生物钟形成后夜醒会自然减少'
        })
    
    patterns_by_type = {'warning': [], 'info': [], 'success': []}
    for p in patterns:
        patterns_by_type.get(p['type'], []).append(p)
    
    for wp in patterns_by_type.get('warning', []):
        advice['general_tips'].append(f"🔴 {wp['title']}：{wp['detail']}")
    for ip in patterns_by_type.get('info', []):
        advice['general_tips'].append(f"🔵 {ip['title']}：{ip['detail']}")
    for sp in patterns_by_type.get('success', []):
        advice['general_tips'].append(f"🟢 {sp['title']}：{sp['detail']}")
    
    if not advice['priority_actions']:
        advice['priority_actions'].append({
            'level': 'low',
            'action': '维持当前良好习惯',
            'detail': '未发现需要紧急调整的问题，继续保持当前的作息规律即可'
        })
    
    return advice


def export_report_to_excel(df, stats, stability, patterns, advice):
    output = BytesIO()
    wb = openpyxl.Workbook()
    
    header_font = Font(name='Arial Unicode MS', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    title_font = Font(name='Arial Unicode MS', bold=True, size=14, color='1F2937')
    section_font = Font(name='Arial Unicode MS', bold=True, size=11, color='6366F1')
    normal_font = Font(name='Arial Unicode MS', size=10)
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    def style_header_row(ws, row, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
    
    ws1 = wb.active
    ws1.title = '概览报告'
    
    ws1.cell(row=1, column=1, value='宝宝睡眠阶段性分析报告').font = Font(name='Arial Unicode MS', bold=True, size=16, color='6366F1')
    ws1.merge_cells('A1:D1')
    ws1.cell(row=2, column=1, value=f'生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}').font = normal_font
    ws1.cell(row=3, column=1, value=f'数据期间：{stats.get("date_range", ["",""])[0]} ~ {stats.get("date_range", ["",""])[1]}')
    ws1.cell(row=3, column=1).font = normal_font
    
    r = 5
    ws1.cell(row=r, column=1, value='一、核心指标汇总').font = section_font
    ws1.merge_cells(f'A{r}:D{r}')
    r += 1
    
    summary_data = [
        ['记录天数', f'{stats.get("days_recorded", 0)} 天', '平均总睡眠', f'{stats.get("avg_total_sleep_hours", 0)} 小时/天'],
        ['平均夜间睡眠', f'{stats.get("avg_night_sleep_hours", 0)} 小时', '平均白天小睡', f'{stats.get("avg_nap_hours", 0)} 小时/天'],
        ['平均夜醒次数', f'{stats.get("avg_nightwakings", 0)} 次/夜', '夜醒天数占比', f'{stats.get("night_waking_days_pct", 0)} %'],
        ['平均入睡时间', stats.get("avg_bedtime", ""), '平均起床时间', stats.get("avg_wakeup", "")],
        ['平均小睡次数', f'{stats.get("avg_naps_count", 0)} 次', '平均日奶量', f'{stats.get("avg_milk_ml", "数据不足")} ml'],
        ['作息稳定度', f'{stability.get("overall", 50)} 分 ({stability.get("level", "")})', '入睡波动', f'±{stability.get("bedtime_std_minutes", 0)} 分钟'],
    ]
    
    for row_data in summary_data:
        for i, val in enumerate(row_data):
            c = ws1.cell(row=r, column=i + 1, value=val)
            c.font = normal_font
            c.alignment = wrap_alignment
            c.border = thin_border
        r += 1
    
    r += 1
    ws1.cell(row=r, column=1, value='二、作息稳定度详情').font = section_font
    ws1.merge_cells(f'A{r}:D{r}')
    r += 1
    
    stability_data = [
        ['维度', '得分', '说明'],
        ['入睡时间稳定性', f'{stability.get("bedtime_stability", 0)} 分', '标准差越小越稳定，理想±30分钟内'],
        ['起床时间稳定性', f'{stability.get("wakeup_stability", 0)} 分', '固定起床时间有助于建立生物钟'],
        ['白天小睡稳定性', f'{stability.get("nap_stability", 0)} 分', '固定小睡时间帮助形成条件反射'],
        ['评估说明', stability.get('details', ''), ''],
    ]
    
    for i, row_data in enumerate(stability_data):
        for j, val in enumerate(row_data):
            c = ws1.cell(row=r, column=j + 1, value=val)
            c.font = normal_font
            c.alignment = wrap_alignment
            c.border = thin_border
        if i == 0:
            style_header_row(ws1, r, len(row_data))
        r += 1
    
    r += 1
    ws1.cell(row=r, column=1, value='三、模式识别结果').font = section_font
    ws1.merge_cells(f'A{r}:D{r}')
    r += 1
    
    pattern_data = [['类型', '标题', '详情说明']]
    for p in patterns:
        type_map = {'warning': '⚠️ 风险', 'info': 'ℹ️ 提示', 'success': '✅ 良好'}
        pattern_data.append([type_map.get(p['type'], p['type']), p['title'], p['detail']])
    
    for i, row_data in enumerate(pattern_data):
        for j, val in enumerate(row_data):
            c = ws1.cell(row=r, column=j + 1, value=val)
            c.font = normal_font
            c.alignment = wrap_alignment
            c.border = thin_border
        if i == 0:
            style_header_row(ws1, r, len(row_data))
        r += 1
    
    r += 1
    ws1.cell(row=r, column=1, value='四、节律调整建议').font = section_font
    ws1.merge_cells(f'A{r}:D{r}')
    r += 1
    
    bw = advice.get('bedtime_window', {})
    if bw:
        cur = bw.get('current', {})
        rec = bw.get('recommended', {})
        ws1.cell(row=r, column=1, value='入睡窗口分析').font = Font(name='Arial Unicode MS', bold=True, size=10)
        r += 1
        ws1.cell(row=r, column=1, value=f'当前平均: {cur.get("avg", "")}，建议窗口: {cur.get("window", "")}').font = normal_font
        r += 1
        ws1.cell(row=r, column=1, value=f'当前状态: {cur.get("consistency", "")}').font = normal_font
        r += 1
        ws1.cell(row=r, column=1, value=f'推荐窗口: {rec.get("window", "")}').font = normal_font
        r += 1
        ws1.cell(row=r, column=1, value=f'推荐理由: {rec.get("reason", "")}').font = normal_font
        r += 1
    
    if advice.get('nap_adjustment'):
        r += 1
        ws1.cell(row=r, column=1, value='小睡调整建议').font = Font(name='Arial Unicode MS', bold=True, size=10)
        r += 1
        for tip in advice['nap_adjustment']:
            ws1.cell(row=r, column=1, value=tip).font = normal_font
            ws1.merge_cells(f'A{r}:D{r}')
            ws1.cell(row=r, column=1).alignment = wrap_alignment
            r += 1
    
    if advice.get('priority_actions'):
        r += 1
        ws1.cell(row=r, column=1, value='优先行动项').font = Font(name='Arial Unicode MS', bold=True, size=10)
        r += 1
        pa_data = [['优先级', '行动', '详情']]
        level_map = {'high': '🔴 高', 'medium': '🟡 中', 'low': '🟢 低'}
        for pa in advice['priority_actions']:
            pa_data.append([level_map.get(pa['level'], pa['level']), pa['action'], pa['detail']])
        for i, row_data in enumerate(pa_data):
            for j, val in enumerate(row_data):
                c = ws1.cell(row=r, column=j + 1, value=val)
                c.font = normal_font
                c.alignment = wrap_alignment
                c.border = thin_border
            if i == 0:
                style_header_row(ws1, r, len(row_data))
            r += 1
    
    if advice.get('general_tips'):
        r += 1
        ws1.cell(row=r, column=1, value='综合提示').font = Font(name='Arial Unicode MS', bold=True, size=10)
        r += 1
        for tip in advice['general_tips']:
            ws1.cell(row=r, column=1, value=tip).font = normal_font
            ws1.merge_cells(f'A{r}:D{r}')
            ws1.cell(row=r, column=1).alignment = wrap_alignment
            r += 1
    
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 28
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 38
    
    ws2 = wb.create_sheet('每日原始数据')
    export_cols = ['date', 'age_months', 'age_group', 'bedtime', 'wakeup_time',
                   'night_sleep_hours', 'nap_hours', 'total_sleep_hours',
                   'nightwakings', 'naps_count', 'milk_amount_ml',
                   'feeding_type', 'teething', 'weather', 'nw_period_group']
    col_names = {
        'date': '日期', 'age_months': '月龄', 'age_group': '月龄阶段',
        'bedtime': '入睡时间', 'wakeup_time': '起床时间',
        'night_sleep_hours': '夜间睡眠(小时)', 'nap_hours': '小睡(小时)',
        'total_sleep_hours': '总睡眠(小时)', 'nightwakings': '夜醒次数',
        'naps_count': '小睡次数', 'milk_amount_ml': '奶量(ml)',
        'feeding_type': '喂养方式', 'teething': '是否长牙',
        'weather': '天气', 'nw_period_group': '主要夜醒时段'
    }
    
    df_export = df.copy()
    df_export['date'] = df_export['date'].dt.strftime('%Y-%m-%d')
    existing_cols = [c for c in export_cols if c in df_export.columns]
    df_export = df_export[existing_cols].rename(columns=col_names)
    
    for i, row in enumerate(dataframe_to_rows(df_export, index=False, header=True)):
        for j, val in enumerate(row):
            c = ws2.cell(row=i + 1, column=j + 1, value=val)
            c.font = normal_font
            c.alignment = wrap_alignment
            c.border = thin_border
        if i == 0:
            style_header_row(ws2, i + 1, len(row))
    
    for col_idx in range(1, len(df_export.columns) + 1):
        ws2.column_dimensions[chr(64 + col_idx)].width = 16
    
    ws3 = wb.create_sheet('分组统计')
    r3 = 1
    
    group_defs = [
        ('按月龄阶段', 'age_group', ['夜间睡眠(小时)', '小睡(小时)', '总睡眠(小时)', '夜醒次数']),
        ('按入睡时段', 'bedtime_group', ['夜间睡眠(小时)', '总睡眠(小时)', '夜醒次数']),
        ('按小睡次数', 'naps_group', ['夜间睡眠(小时)', '总睡眠(小时)', '夜醒次数']),
        ('按奶量区间', 'milk_group', ['夜间睡眠(小时)', '总睡眠(小时)', '夜醒次数']),
        ('按喂养方式', 'feeding_type', ['夜间睡眠(小时)', '总睡眠(小时)', '夜醒次数']),
    ]
    
    eng_to_cn = {
        'night_sleep_hours': '夜间睡眠(小时)', 'nap_hours': '小睡(小时)',
        'total_sleep_hours': '总睡眠(小时)', 'nightwakings': '夜醒次数'
    }
    
    for title, group_col, targets in group_defs:
        cn_targets = [eng_to_cn.get(t, t) for t in targets]
        if group_col not in df.columns:
            continue
        df_tmp = df.dropna(subset=[group_col]).copy()
        if len(df_tmp) < 2:
            continue
        grouped = df_tmp.groupby(group_col)[targets].agg(['mean', 'count']).round(2)
        grouped.columns = [f'{eng_to_cn.get(c, c)}-{s}' for c, s in grouped.columns]
        grouped = grouped.reset_index()
        
        ws3.cell(row=r3, column=1, value=title).font = section_font
        ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=len(grouped.columns))
        r3 += 1
        
        for j, col_name in enumerate(grouped.columns):
            col_display = col_name if j > 0 else '分组'
            c = ws3.cell(row=r3, column=j + 1, value=col_display)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border
        r3 += 1
        
        for _, row_data in grouped.iterrows():
            for j, val in enumerate(row_data):
                c = ws3.cell(row=r3, column=j + 1, value=val)
                c.font = normal_font
                c.alignment = wrap_alignment
                c.border = thin_border
            r3 += 1
        r3 += 2
    
    for col_idx in range(1, 15):
        ws3.column_dimensions[chr(64 + col_idx)].width = 18
    
    wb.save(output)
    output.seek(0)
    return output
