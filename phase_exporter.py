import pandas as pd
import numpy as np
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from data_processor import minutes_to_time_display


def export_phase_review_to_excel(
    phase_results, comparison, summary,
    df_phases_dict, intervention_comparison=None,
    quality_result=None, filters_info=None
):
    output = BytesIO()
    wb = openpyxl.Workbook()
    
    header_font = Font(name='Arial Unicode MS', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    section_font = Font(name='Arial Unicode MS', bold=True, size=11, color='6366F1')
    sub_section_font = Font(name='Arial Unicode MS', bold=True, size=10, color='374151')
    normal_font = Font(name='Arial Unicode MS', size=10)
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    status_fill_map = {
        '改善中': PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid'),
        '改善中（稳定性待提升）': PatternFill(start_color='A7F3D0', end_color='A7F3D0', fill_type='solid'),
        '稳定良好': PatternFill(start_color='6EE7B7', end_color='6EE7B7', fill_type='solid'),
        '反复波动': PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid'),
        '观察中': PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid'),
        '不稳定': PatternFill(start_color='FED7AA', end_color='FED7AA', fill_type='solid'),
        '阶段倒退': PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'),
        '数据不足': PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid'),
    }
    
    def style_header_row(ws, row, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
    
    def style_data_cell(ws, row, col, value, fill=None):
        if value is None:
            value = ''
        elif isinstance(value, (float, np.floating)):
            if pd.isna(value) or np.isnan(value) or np.isinf(value):
                value = ''
        elif pd.isna(value):
            value = ''
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = normal_font
        cell.alignment = wrap_alignment
        cell.border = thin_border
        if fill:
            cell.fill = fill
        return cell
    
    ws1 = wb.active
    ws1.title = '阶段复盘概览'
    
    ws1.cell(row=1, column=1, value='宝宝睡眠阶段复盘报告').font = Font(
        name='Arial Unicode MS', bold=True, size=16, color='6366F1'
    )
    ws1.merge_cells('A1:F1')
    ws1.cell(row=2, column=1, value=f'生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}').font = normal_font
    
    r = 4
    ws1.cell(row=r, column=1, value='一、阶段定义').font = section_font
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    
    phase_def_headers = ['阶段名称', '开始日期', '结束日期', '记录天数', '阶段状态', '说明']
    for j, h in enumerate(phase_def_headers):
        style_data_cell(ws1, r, j + 1, h)
    style_header_row(ws1, r, len(phase_def_headers))
    r += 1
    
    for pr in phase_results:
        dr = pr.get('date_range')
        start_d = dr[0] if dr else '-'
        end_d = dr[1] if dr else '-'
        status = pr.get('status', '数据不足')
        fill = status_fill_map.get(status)
        
        style_data_cell(ws1, r, 1, pr['phase_name'])
        style_data_cell(ws1, r, 2, start_d)
        style_data_cell(ws1, r, 3, end_d)
        style_data_cell(ws1, r, 4, f"{pr.get('days_count', 0)} 天")
        style_data_cell(ws1, r, 5, status, fill=fill)
        
        notes = []
        if status == '改善中':
            notes.append('夜醒呈下降趋势')
        elif status == '阶段倒退':
            notes.append('需关注退步原因')
        elif status == '反复波动':
            notes.append('数据起伏大')
        elif status == '稳定良好':
            notes.append('睡眠规律已形成')
        style_data_cell(ws1, r, 6, '；'.join(notes) if notes else '-')
        r += 1
    
    if filters_info:
        r += 1
        ws1.cell(row=r, column=1, value='筛选条件').font = section_font
        ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
        for k, v in filters_info.items():
            style_data_cell(ws1, r, 1, k)
            style_data_cell(ws1, r, 2, str(v))
            ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            r += 1
    
    r += 1
    ws1.cell(row=r, column=1, value='二、核心指标对比').font = section_font
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    
    metric_defs = [
        ('avg_total_sleep_hours', '平均总睡眠(小时/天)'),
        ('avg_night_sleep_hours', '平均夜间睡眠(小时)'),
        ('avg_nap_hours', '平均白天小睡(小时)'),
        ('avg_nightwakings', '平均夜醒(次/夜)'),
        ('night_waking_days_pct', '夜醒天数占比(%)'),
        ('avg_bedtime', '平均入睡时间'),
        ('avg_wakeup', '平均起床时间'),
        ('avg_naps_count', '平均小睡次数'),
        ('avg_milk_ml', '平均奶量(ml)'),
        ('avg_last_nap_end', '平均最后一觉结束时间'),
        ('bedtime_window_stability', '入睡窗口稳定度(分)'),
    ]
    
    headers = ['指标'] + [pr['phase_name'] for pr in phase_results]
    if len(phase_results) >= 2:
        headers.append('变化(后-前)')
        headers.append('变化率(%)')
    
    for j, h in enumerate(headers):
        style_data_cell(ws1, r, j + 1, h)
    style_header_row(ws1, r, len(headers))
    r += 1
    
    for metric_key, metric_label in metric_defs:
        style_data_cell(ws1, r, 1, metric_label)
        
        values = []
        for i, pr in enumerate(phase_results):
            val = pr['metrics'].get(metric_key)
            if isinstance(val, float):
                display_val = round(val, 2)
            else:
                display_val = val if val else '-'
            values.append(val if isinstance(val, (int, float)) else None)
            style_data_cell(ws1, r, i + 2, display_val)
        
        if len(values) >= 2 and all(v is not None for v in values[:2]):
            try:
                diff = float(values[-1]) - float(values[0])
                base_val = abs(float(values[0]))
                if base_val > 0.001:
                    pct = (diff / base_val) * 100
                else:
                    pct = 0.0 if abs(diff) < 0.001 else (100.0 if diff > 0 else -100.0)
                style_data_cell(ws1, r, len(phase_results) + 2, round(diff, 2))
                style_data_cell(ws1, r, len(phase_results) + 3, f"{round(pct, 1)}%")
            except (TypeError, ValueError):
                style_data_cell(ws1, r, len(phase_results) + 2, '-')
                style_data_cell(ws1, r, len(phase_results) + 3, '-')
        
        r += 1
    
    r += 1
    ws1.cell(row=r, column=1, value='三、作息稳定度对比').font = section_font
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    
    stab_headers = ['维度'] + [pr['phase_name'] for pr in phase_results]
    for j, h in enumerate(stab_headers):
        style_data_cell(ws1, r, j + 1, h)
    style_header_row(ws1, r, len(stab_headers))
    r += 1
    
    stab_metrics = [
        ('overall', '综合稳定度(分)'),
        ('bedtime_stability', '入睡时间稳定(分)'),
        ('wakeup_stability', '起床时间稳定(分)'),
        ('nap_stability', '小睡稳定(分)'),
    ]
    for key, label in stab_metrics:
        style_data_cell(ws1, r, 1, label)
        for i, pr in enumerate(phase_results):
            val = pr['stability'].get(key, 0)
            style_data_cell(ws1, r, i + 2, round(val, 1))
        r += 1
    
    r += 1
    ws1.cell(row=r, column=1, value='四、夜醒时段分布(次数)').font = section_font
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    
    nw_periods = ['入睡后(22-01)', '深夜(01-04)', '凌晨(04-06)', '清晨(06+)']
    nw_headers = ['夜醒时段'] + [pr['phase_name'] for pr in phase_results]
    for j, h in enumerate(nw_headers):
        style_data_cell(ws1, r, j + 1, h)
    style_header_row(ws1, r, len(nw_headers))
    r += 1
    
    for period in nw_periods:
        style_data_cell(ws1, r, 1, period)
        for i, pr in enumerate(phase_results):
            count = pr.get('nw_period_distribution', {}).get(period, 0)
            style_data_cell(ws1, r, i + 2, count)
        r += 1
    
    r += 1
    ws1.cell(row=r, column=1, value='五、奶量与夜醒关联').font = section_font
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    
    milk_headers = ['阶段', '相关系数', '说明']
    for j, h in enumerate(milk_headers):
        style_data_cell(ws1, r, j + 1, h)
    style_header_row(ws1, r, len(milk_headers))
    r += 1
    
    for pr in phase_results:
        corr = pr.get('milk_nw_correlation')
        style_data_cell(ws1, r, 1, pr['phase_name'])
        if corr is not None and isinstance(corr, (int, float)) and pd.notna(corr):
            style_data_cell(ws1, r, 2, f'{float(corr):.3f}')
        else:
            style_data_cell(ws1, r, 2, '数据不足')
        note = ''
        if corr is not None and isinstance(corr, (int, float)) and pd.notna(corr):
            corr_val = float(corr)
            if corr_val < -0.3:
                note = '奶量增加显著关联夜醒减少'
            elif corr_val < -0.1:
                note = '奶量增加轻度关联夜醒减少'
            elif corr_val <= 0.1:
                note = '无明显关联'
            else:
                note = '奶量增加反而夜醒增多，需排查原因'
        style_data_cell(ws1, r, 3, note)
        r += 1
    
    if intervention_comparison:
        r += 1
        ws1.cell(row=r, column=1, value='六、干预实际效果与预测偏差').font = section_font
        ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
        
        inter_headers = ['指标', '基线值', '预测值', '实际值', '与预测偏差', '目标达成率(%)']
        for j, h in enumerate(inter_headers):
            style_data_cell(ws1, r, j + 1, h)
        style_header_row(ws1, r, len(inter_headers))
        r += 1
        
        inter_metric_labels = {
            'nightwakings': '夜醒次数(次/夜)',
            'total_sleep_hours': '总睡眠时长(小时/天)',
            'stability_score': '作息稳定度(分)',
        }
        for key, label in inter_metric_labels.items():
            data = intervention_comparison.get(key, {})
            style_data_cell(ws1, r, 1, label)
            style_data_cell(ws1, r, 2, data.get('baseline', 0))
            style_data_cell(ws1, r, 3, data.get('predicted', 0))
            style_data_cell(ws1, r, 4, data.get('actual', 0))
            style_data_cell(ws1, r, 5, data.get('diff_pred_actual', 0))
            style_data_cell(ws1, r, 6, f"{data.get('achievement_pct', 0)}%")
            r += 1
    
    ws1.column_dimensions['A'].width = 26
    for col_idx in range(2, 7):
        ws1.column_dimensions[chr(64 + col_idx)].width = 20
    
    ws2 = wb.create_sheet('趋势对比数据')
    r2 = 1
    
    ws2.cell(row=r2, column=1, value='各阶段每日数据（用于趋势图）').font = section_font
    ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=10)
    r2 += 2
    
    for phase_name, df_phase in df_phases_dict.items():
        ws2.cell(row=r2, column=1, value=phase_name).font = sub_section_font
        r2 += 1
        
        trend_cols = ['date', 'total_sleep_hours', 'night_sleep_hours', 'nap_hours',
                      'nightwakings', 'naps_count', 'milk_amount_ml', 'bedtime', 'wakeup_time']
        col_labels = {
            'date': '日期', 'total_sleep_hours': '总睡眠(小时)',
            'night_sleep_hours': '夜间睡眠(小时)', 'nap_hours': '小睡(小时)',
            'nightwakings': '夜醒次数', 'naps_count': '小睡次数',
            'milk_amount_ml': '奶量(ml)', 'bedtime': '入睡时间', 'wakeup_time': '起床时间'
        }
        existing = [c for c in trend_cols if c in df_phase.columns]
        
        headers = ['阶段内天数'] + [col_labels.get(c, c) for c in existing]
        for j, h in enumerate(headers):
            style_data_cell(ws2, r2, j + 1, h)
        style_header_row(ws2, r2, len(headers))
        r2 += 1
        
        df_sorted = df_phase.sort_values('date').reset_index(drop=True)
        for idx, row in df_sorted.iterrows():
            style_data_cell(ws2, r2, 1, idx + 1)
            for j, col in enumerate(existing):
                val = row[col]
                if col == 'date':
                    val = val.strftime('%Y-%m-%d') if pd.notna(val) else ''
                elif isinstance(val, float) or isinstance(val, np.floating):
                    if pd.isna(val):
                        val = ''
                    else:
                        val = round(float(val), 2)
                elif pd.isna(val):
                    val = ''
                style_data_cell(ws2, r2, j + 2, val)
            r2 += 1
        r2 += 2
    
    for col_idx in range(1, 12):
        ws2.column_dimensions[chr(64 + col_idx)].width = 16
    
    ws3 = wb.create_sheet('关键变化与建议')
    r3 = 1
    
    ws3.cell(row=r3, column=1, value='关键变化摘要与下一阶段建议').font = Font(
        name='Arial Unicode MS', bold=True, size=14, color='6366F1'
    )
    ws3.merge_cells('A1:C1')
    r3 += 2
    
    if summary:
        ws3.cell(row=r3, column=1, value='一、整体评估').font = section_font
        ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=3)
        r3 += 1
        style_data_cell(ws3, r3, 1, summary.get('overall_assessment', ''))
        ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=3)
        r3 += 2
        
        ws3.cell(row=r3, column=1, value='二、关键变化').font = section_font
        ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=3)
        r3 += 1
        for change in summary.get('key_changes', []):
            style_data_cell(ws3, r3, 1, f'• {change}')
            ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=3)
            r3 += 1
        r3 += 1
        
        ws3.cell(row=r3, column=1, value='三、下一阶段建议').font = section_font
        ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=3)
        r3 += 1
        for i, rec in enumerate(summary.get('recommendations', [])):
            style_data_cell(ws3, r3, 1, f'{i+1}. {rec}')
            ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=3)
            r3 += 1
    
    if comparison:
        r3 += 1
        ws3.cell(row=r3, column=1, value='四、详细指标变化').font = section_font
        ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=3)
        r3 += 1
        
        metric_labels = {
            'avg_total_sleep_hours': '总睡眠(小时/天)',
            'avg_night_sleep_hours': '夜间睡眠(小时)',
            'avg_nap_hours': '白天小睡(小时)',
            'avg_nightwakings': '夜醒次数(次/夜)',
            'bedtime_window_stability': '入睡窗口稳定度(分)',
            'stability_score': '综合稳定度(分)',
            'avg_milk_ml': '平均奶量(ml)',
            'avg_naps_count': '小睡次数',
            'night_waking_days_pct': '夜醒天数占比(%)',
            'last_nap_minutes': '最后一觉结束时间',
        }
        
        for comp_label, changes in comparison.items():
            ws3.cell(row=r3, column=1, value=comp_label).font = sub_section_font
            r3 += 1
            
            headers = ['指标', '基准值', '对比值', '差值', '变化率(%)']
            for j, h in enumerate(headers):
                style_data_cell(ws3, r3, j + 1, h)
            style_header_row(ws3, r3, len(headers))
            r3 += 1
            
            for key, data in changes.items():
                style_data_cell(ws3, r3, 1, metric_labels.get(key, key))
                style_data_cell(ws3, r3, 2, data.get('base', ''))
                style_data_cell(ws3, r3, 3, data.get('compare', ''))
                style_data_cell(ws3, r3, 4, data.get('diff', ''))
                style_data_cell(ws3, r3, 5, f"{data.get('pct', '')}%")
                r3 += 1
            r3 += 1
    
    r3 += 1
    ws3.cell(row=r3, column=1, value='五、异常记录影响说明').font = section_font
    ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=3)
    r3 += 1
    
    if quality_result:
        score = quality_result.get('score', 0)
        level = quality_result.get('level', '')
        total = quality_result.get('total_records', 0)
        valid = quality_result.get('valid_records', 0)
        excluded = quality_result.get('excluded_records', 0)
        
        quality_items = [
            ('数据质量评分', f'{score} 分 ({level})'),
            ('总记录数', f'{total} 条'),
            ('有效记录(纳入统计)', f'{valid} 条'),
            ('排除记录(影响睡眠时长)', f'{excluded} 条'),
        ]
        for label, val in quality_items:
            style_data_cell(ws3, r3, 1, label)
            style_data_cell(ws3, r3, 2, val)
            ws3.merge_cells(start_row=r3, start_column=2, end_row=r3, end_column=3)
            r3 += 1
        
        anomaly_records = quality_result.get('anomaly_records', [])
        if anomaly_records:
            r3 += 1
            ws3.cell(row=r3, column=1, value=f'异常记录明细（共 {len(anomaly_records)} 条）').font = sub_section_font
            r3 += 1
            
            anom_headers = ['日期', '严重程度', '问题描述', '影响字段']
            for j, h in enumerate(anom_headers):
                style_data_cell(ws3, r3, j + 1, h)
            style_header_row(ws3, r3, len(anom_headers))
            r3 += 1
            
            for rec in anomaly_records[:50]:
                try:
                    issues_list = rec.get('issues', [])
                    if issues_list and isinstance(issues_list, list):
                        issues_str = '；'.join([
                            i.get('message', str(i)) if isinstance(i, dict) else str(i) 
                            for i in issues_list
                        ])
                    else:
                        issues_str = str(issues_list) if issues_list else '无'
                    
                    affected_fields = rec.get('affected_fields', [])
                    if isinstance(affected_fields, list):
                        affected = '、'.join(affected_fields) or '未知'
                    else:
                        affected = str(affected_fields)
                    
                    severity = rec.get('severity', 'unknown')
                    sev_fill = status_fill_map.get('阶段倒退') if severity == 'critical' else status_fill_map.get('反复波动')
                    
                    style_data_cell(ws3, r3, 1, rec.get('date', ''), fill=sev_fill)
                    style_data_cell(ws3, r3, 2, severity, fill=sev_fill)
                    style_data_cell(ws3, r3, 3, issues_str, fill=sev_fill)
                    style_data_cell(ws3, r3, 4, affected, fill=sev_fill)
                except Exception as e:
                    style_data_cell(ws3, r3, 1, rec.get('date', '未知日期'))
                    style_data_cell(ws3, r3, 2, 'unknown')
                    style_data_cell(ws3, r3, 3, f'记录解析异常: {str(e)}')
                    style_data_cell(ws3, r3, 4, '未知')
                r3 += 1
            
            if len(anomaly_records) > 50:
                style_data_cell(ws3, r3, 1, f'... 还有 {len(anomaly_records) - 50} 条异常记录未显示')
                ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=4)
                r3 += 1
    else:
        style_data_cell(ws3, r3, 1, '无数据质量评估结果')
        ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=3)
        r3 += 1
    
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 22
    ws3.column_dimensions['C'].width = 22
    
    wb.save(output)
    output.seek(0)
    return output
