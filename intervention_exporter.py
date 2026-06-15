import pandas as pd
import numpy as np
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from data_processor import minutes_to_time_display
from intervention_params import get_param_display


def export_intervention_plan_to_excel(
    filtered_df, baseline, params, prediction, daily_plan, calendar_df,
    priorities, risks, execution_summary, baseline_summary, param_display
):
    output = BytesIO()
    wb = openpyxl.Workbook()
    
    header_font = Font(name='Arial Unicode MS', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    title_font = Font(name='Arial Unicode MS', bold=True, size=14, color='1F2937')
    section_font = Font(name='Arial Unicode MS', bold=True, size=11, color='6366F1')
    sub_section_font = Font(name='Arial Unicode MS', bold=True, size=10, color='374151')
    normal_font = Font(name='Arial Unicode MS', size=10)
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    risk_fill_map = {
        'high': PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'),
        'medium': PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid'),
        'low': PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid'),
    }
    
    priority_fill_map = {
        'high': PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'),
        'medium': PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid'),
        'low': PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid'),
    }
    
    def style_header_row(ws, row, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
    
    def style_data_cell(ws, row, col, value, fill=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = normal_font
        cell.alignment = wrap_alignment
        cell.border = thin_border
        if fill:
            cell.fill = fill
        return cell
    
    ws1 = wb.active
    ws1.title = '干预计划概览'
    
    ws1.cell(row=1, column=1, value='宝宝睡眠干预方案报告').font = Font(
        name='Arial Unicode MS', bold=True, size=16, color='6366F1'
    )
    ws1.merge_cells('A1:D1')
    ws1.cell(row=2, column=1, value=f'生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}').font = normal_font
    ws1.cell(row=3, column=1, value=f'方案名称：{execution_summary.get("intervention_name", "综合睡眠改善方案")}').font = normal_font
    ws1.cell(row=4, column=1, value=f'干预周期：{execution_summary.get("duration", "14 天")}').font = normal_font
    
    r = 6
    ws1.cell(row=r, column=1, value='一、基线摘要').font = section_font
    ws1.merge_cells(f'A{r}:D{r}')
    r += 1
    
    baseline_items = [
        ('记录天数', baseline_summary.get('记录天数', '-')),
        ('平均总睡眠', baseline_summary.get('平均总睡眠', '-')),
        ('平均夜间睡眠', baseline_summary.get('平均夜间睡眠', '-')),
        ('平均夜醒次数', baseline_summary.get('平均夜醒次数', '-')),
        ('夜醒天数占比', baseline_summary.get('夜醒天数占比', '-')),
        ('平均小睡次数', baseline_summary.get('平均小睡次数', '-')),
        ('平均奶量', baseline_summary.get('平均奶量', '-')),
        ('作息稳定度', baseline_summary.get('作息稳定度', '-')),
        ('平均入睡时间', baseline_summary.get('平均入睡时间', '-')),
        ('平均起床时间', baseline_summary.get('平均起床时间', '-')),
    ]
    
    for label, value in baseline_items:
        style_data_cell(ws1, r, 1, label)
        style_data_cell(ws1, r, 2, value)
        style_data_cell(ws1, r, 3, '')
        style_data_cell(ws1, r, 4, '')
        r += 1
    
    r += 1
    ws1.cell(row=r, column=1, value='二、干预参数').font = section_font
    ws1.merge_cells(f'A{r}:D{r}')
    r += 1
    
    param_items = [
        ('目标入睡窗口', param_display['bedtime_window']),
        ('最后一觉最晚结束', param_display['last_nap_deadline']),
        ('白天小睡调整', param_display['nap_adjustment']),
        ('奶量变化', param_display['milk_change']),
        ('夜醒安抚策略', param_display['soothing_strategy']),
        ('模拟周期', param_display['sim_duration']),
    ]
    
    for label, value in param_items:
        style_data_cell(ws1, r, 1, label)
        style_data_cell(ws1, r, 2, value)
        style_data_cell(ws1, r, 3, '')
        style_data_cell(ws1, r, 4, '')
        r += 1
    
    r += 1
    ws1.cell(row=r, column=1, value='三、预测结果对比').font = section_font
    ws1.merge_cells(f'A{r}:D{r}')
    r += 1
    
    pred_data = [
        ['指标', '基线值', '预测值', '变化'],
        ['夜醒次数(次/夜)', 
         f"{baseline.get('avg_nightwakings', 0):.2f}",
         f"{prediction['predicted']['avg_nightwakings']:.2f}",
         f"减少 {prediction['changes']['night_waking_reduction_pct']:.1f}%"],
        ['总睡眠时长(小时/天)', 
         f"{baseline.get('avg_total_sleep_hours', 0):.1f}",
         f"{prediction['predicted']['avg_total_sleep_hours']:.1f}",
         f"{'增加' if prediction['changes']['total_sleep_change_minutes'] >= 0 else '减少'} {abs(prediction['changes']['total_sleep_change_minutes']):.0f} 分钟"],
        ['作息稳定度(分)', 
         f"{baseline.get('stability_score', 0):.1f}",
         f"{prediction['predicted']['stability_score']:.1f}",
         f"提升 {prediction['changes']['stability_gain']:.1f} 分"],
    ]
    
    for i, row_data in enumerate(pred_data):
        for j, val in enumerate(row_data):
            style_data_cell(ws1, r, j + 1, val)
        if i == 0:
            style_header_row(ws1, r, len(row_data))
        r += 1
    
    r += 1
    ws1.cell(row=r, column=1, value='四、各维度影响分析').font = section_font
    ws1.merge_cells(f'A{r}:D{r}')
    r += 1
    
    dim_data = [['维度', '夜醒改善贡献', '睡眠时长贡献', '风险等级']]
    for key, dim in prediction.get('dimension_effects', {}).items():
        risk_label = {'low': '低', 'medium': '中', 'high': '高'}.get(dim['risk_level'], '低')
        dim_data.append([
            f"{dim['icon']} {dim['name']}",
            f"{dim['nw_reduction_pct_contribution']:+.1f}%",
            f"{dim['sleep_change_contribution']:+.0f} 分钟",
            risk_label,
        ])
    
    for i, row_data in enumerate(dim_data):
        for j, val in enumerate(row_data):
            style_data_cell(ws1, r, j + 1, val)
        if i == 0:
            style_header_row(ws1, r, len(row_data))
        r += 1
    
    r += 1
    ws1.cell(row=r, column=1, value='五、执行优先级').font = section_font
    ws1.merge_cells(f'A{r}:D{r}')
    r += 1
    
    priority_data = [['优先级', '维度', '夜醒影响', '风险']]
    for p in priorities:
        priority_label = {'high': '🔴 最高', 'medium': '🟡 中等', 'low': '🟢 较低'}.get(p['priority'], '低')
        risk_label = {'low': '低', 'medium': '中', 'high': '高'}.get(p['risk_level'], '低')
        priority_data.append([
            priority_label,
            f"{p['icon']} {p['name']}",
            f"{p['night_waking_impact_pct']:.1f}%",
            risk_label,
        ])
    
    for i, row_data in enumerate(priority_data):
        fill = None
        if i > 0:
            p = priorities[i - 1]
            fill = priority_fill_map.get(p['priority'])
        for j, val in enumerate(row_data):
            style_data_cell(ws1, r, j + 1, val, fill=fill)
        if i == 0:
            style_header_row(ws1, r, len(row_data))
        r += 1
    
    r += 1
    ws1.cell(row=r, column=1, value='六、风险提示').font = section_font
    ws1.merge_cells(f'A{r}:D{r}')
    r += 1
    
    risk_data = [['风险等级', '标题', '详情说明']]
    for risk in risks:
        level_label = {'high': '🔴 高', 'medium': '🟡 中', 'low': '🟢 低'}.get(risk['level'], '低')
        risk_data.append([level_label, risk['title'], risk['detail']])
    
    for i, row_data in enumerate(risk_data):
        fill = None
        if i > 0:
            risk = risks[i - 1]
            fill = risk_fill_map.get(risk['level'])
        for j, val in enumerate(row_data):
            style_data_cell(ws1, r, j + 1, val, fill=fill)
        if i == 0:
            style_header_row(ws1, r, len(row_data))
        r += 1
    
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 38
    
    ws2 = wb.create_sheet('干预日历表')
    
    ws2.cell(row=1, column=1, value='睡眠干预执行日历').font = Font(
        name='Arial Unicode MS', bold=True, size=14, color='6366F1'
    )
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws2.cell(row=2, column=1, value=f'生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}').font = normal_font
    
    cal_df = calendar_df.copy()
    weekday_map = {'Monday': '周一', 'Tuesday': '周二', 'Wednesday': '周三',
                 'Thursday': '周四', 'Friday': '周五', 'Saturday': '周六', 'Sunday': '周日'}
    cal_df['星期'] = cal_df['天数'].apply(lambda x: '')
    
    display_cols = ['日期', '天数', '阶段', '入睡窗口', '最后一觉截止', 
                   '小睡目标', '奶量目标', '安抚策略', '预计夜醒', '关键任务']
    cal_df = cal_df[display_cols]
    
    for i, row in enumerate(dataframe_to_rows(cal_df, index=False, header=True)):
        row_num = i + 4
        phase = ''
        if i > 0:
            phase = cal_df.iloc[i-1]['阶段']
        for j, val in enumerate(row):
            fill = None
            if phase == '适应期':
                fill = PatternFill(start_color='F0F9FF', end_color='F0F9FF', fill_type='solid')
            elif phase == '调整期':
                fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            elif phase == '巩固期':
                fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
            elif phase == '稳定期':
                fill = PatternFill(start_color='E0E7FF', end_color='E0E7FF', fill_type='solid')
            style_data_cell(ws2, row_num, j + 1, val, fill=fill if i > 0 else None)
        if i == 0:
            style_header_row(ws2, row_num, len(row))
    
    col_widths = [12, 10, 10, 18, 14, 12, 12, 14, 12, 40]
    for i, w in enumerate(col_widths):
        ws2.column_dimensions[chr(65 + i)].width = w
    
    ws3 = wb.create_sheet('每日执行建议')
    
    ws3.cell(row=1, column=1, value='每日详细执行建议').font = Font(
        name='Arial Unicode MS', bold=True, size=14, color='6366F1'
    )
    ws3.merge_cells('A1:D1')
    
    r3 = 3
    
    for day_info in daily_plan:
        day_num = day_info['day']
        phase = day_info['phase']
        
        phase_fill = {
            '适应期': PatternFill(start_color='F0F9FF', end_color='F0F9FF', fill_type='solid'),
            '调整期': PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid'),
            '巩固期': PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid'),
            '稳定期': PatternFill(start_color='E0E7FF', end_color='E0E7FF', fill_type='solid'),
        }.get(phase)
        
        ws3.cell(row=r3, column=1, value=f"第 {day_num} 天 - {phase}").font = sub_section_font
        ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=4)
        if phase_fill:
            for c in range(1, 5):
                ws3.cell(row=r3, column=c).fill = phase_fill
        r3 += 1
        
        detail_items = [
            ('日期', day_info['date']),
            ('目标入睡窗口', day_info['target_bedtime_window']),
            ('最后一觉截止', day_info['last_nap_deadline']),
            ('小睡目标', f"{day_info['target_naps_count']} 次"),
            ('奶量目标', f"{day_info['target_milk_ml']:.0f} ml" if day_info['target_milk_ml'] > 0 else '保持不变'),
            ('安抚策略', day_info['soothing_level']),
            ('预计夜醒', f"{day_info['expected_nightwakings']:.1f} 次"),
        ]
        
        for label, value in detail_items:
            style_data_cell(ws3, r3, 1, label)
            style_data_cell(ws3, r3, 2, value)
            style_data_cell(ws3, r3, 3, '')
            style_data_cell(ws3, r3, 4, '')
            r3 += 1
        
        ws3.cell(row=r3, column=1, value='关键任务').font = Font(name='Arial Unicode MS', bold=True, size=10)
        r3 += 1
        for task in day_info['key_tasks']:
            style_data_cell(ws3, r3, 1, f'• {task}')
            ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=4)
            r3 += 1
        
        if day_info.get('note'):
            style_data_cell(ws3, r3, 1, f"💡 {day_info['note']}")
            ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=4)
            r3 += 1
        
        r3 += 1
    
    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 25
    ws3.column_dimensions['C'].width = 20
    ws3.column_dimensions['D'].width = 35
    
    ws4 = wb.create_sheet('风险与注意事项')
    
    ws4.cell(row=1, column=1, value='风险提示与注意事项').font = Font(
        name='Arial Unicode MS', bold=True, size=14, color='6366F1'
    )
    ws4.merge_cells('A1:C1')
    
    r4 = 3
    ws4.cell(row=r4, column=1, value='一、风险提示').font = section_font
    ws4.merge_cells(f'A{r4}:C{r4}')
    r4 += 1
    
    risk_header = ['风险等级', '标题', '详情说明']
    for j, h in enumerate(risk_header):
        style_data_cell(ws4, r4, j + 1, h)
    style_header_row(ws4, r4, len(risk_header))
    r4 += 1
    
    for risk in risks:
        level_label = {'high': '🔴 高风险', 'medium': '🟡 中风险', 'low': '🟢 低风险'}.get(risk['level'], '低')
        fill = risk_fill_map.get(risk['level'])
        style_data_cell(ws4, r4, 1, level_label, fill=fill)
        style_data_cell(ws4, r4, 2, risk['title'], fill=fill)
        style_data_cell(ws4, r4, 3, risk['detail'], fill=fill)
        r4 += 1
    
    r4 += 2
    ws4.cell(row=r4, column=1, value='二、执行注意事项').font = section_font
    ws4.merge_cells(f'A{r4}:C{r4}')
    r4 += 1
    
    general_notes = [
        '干预效果因人而异，本预测基于历史数据推算，实际效果可能有差异。',
        '建议每天固定时间记录睡眠数据，每周评估调整方案。',
        '如遇宝宝生病、长牙等特殊情况，可适当灵活调整，不必强求。',
        '睡前程序是关键：固定的睡前程序比具体时间更重要。',
        '夜醒安抚需坚持，通常3-7天可见初步效果，2-4周形成稳定习惯。',
        '家长的情绪状态也很重要，保持耐心和一致性是成功的关键。',
    ]
    
    for i, note in enumerate(general_notes):
        style_data_cell(ws4, r4, 1, f'{i+1}.')
        style_data_cell(ws4, r4, 2, note)
        ws4.merge_cells(start_row=r4, start_column=2, end_row=r4, end_column=3)
        r4 += 1
    
    r4 += 2
    ws4.cell(row=r4, column=1, value='三、紧急情况处理').font = section_font
    ws4.merge_cells(f'A{r4}:C{r4}')
    r4 += 1
    
    emergency_notes = [
        ('宝宝发热（>38.5℃）', '暂停训练，优先照顾宝宝健康。'),
        ('严重腹泻或呕吐', '及时就医，待恢复后再继续。'),
        ('夜醒突然大幅增加且难以安抚', '检查是否有身体不适，必要时就医。'),
        ('持续一周以上无改善', '考虑调整方案或咨询专业人士。'),
    ]
    
    for situation, advice in emergency_notes:
        style_data_cell(ws4, r4, 1, situation)
        style_data_cell(ws4, r4, 2, advice)
        ws4.merge_cells(start_row=r4, start_column=2, end_row=r4, end_column=3)
        r4 += 1
    
    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 30
    ws4.column_dimensions['C'].width = 40
    
    wb.save(output)
    output.seek(0)
    return output
